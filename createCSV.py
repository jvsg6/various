# -*- coding: utf-8 -*-
#Модуль для переделки координат и сохранения в csv формате
import sys
import os
import xml.etree.ElementTree as etree
from math import cos,sin,radians,sqrt,fabs,acos,pi,degrees,floor,log10
import xml.etree.ElementTree
import matplotlib.pyplot as plt
import openpyxl
from openpyxl import load_workbook
import numpy as np
path="/home/egor/quest/Mayak/table.xlsx"

def main():
	name=path
	wb = load_workbook(name)
	ws = wb.active
	f = open('100km_Mayak.csv', 'wt')
	for j in range(2, 183):
		for i in range(1, 4):
			if i==1:
				f.write(ws.cell(row=j, column=i).value.encode('utf8'))
				f.write(";")
			if i==2:
				coord=str(ws.cell(row=j, column=i).value.encode('utf8'))
				hour=coord.split('°')[0]
				minute=coord.split('°')[1].split('\'')[0]
				seconds = coord.split('\'')[1].split('\"')[0]
				#print hour, minute, seconds
				whole=float(hour)
				dec=float(minute)/60.0
				decdec=float(seconds)/3600.0
				lat=whole+dec+decdec
				f.write(str("%.6f")%lat)
				#print lat
				f.write(";")
			if i==3:
				coord=str(ws.cell(row=j, column=i).value.encode('utf8'))
				hour=coord.split('°')[0]
				minute=coord.split('°')[1].split('\'')[0]
				seconds = coord.split('\'')[1].split('\"')[0]
				#print hour, minute, seconds
				whole=float(hour)
				dec=float(minute)/60.0
				decdec=float(seconds)/3600.0
				lon=whole+dec+decdec
				f.write(str("%.6f")%lon)
				f.write(";")
				f.write('\n')
			
			
	#f.write("DSAA" + '\n')
	#f.write('\n')	
	f.close()
	
	return 



if __name__ == "__main__":
	sys.exit(main())